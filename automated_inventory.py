from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

import os
from dotenv import load_dotenv
import time

# Initializing the driver, webpage link, login email, and password
def initializing():
    load_dotenv()
    service = Service(executable_path=os.getenv('DRIVER_EXECUTABLE_PATH'))
    driver = webdriver.Chrome(service=service)
    driver.maximize_window()
    driver.get(os.getenv('WEB_LINK'))
    LOGIN_EMAIL = os.getenv('LOGIN_EMAIL')
    LOGIN_PASSWORD = os.getenv('LOGIN_PASSWORD')

    return driver, LOGIN_EMAIL, LOGIN_PASSWORD

# Logging in to the webpage
def login(driver, LOGIN_EMAIL, LOGIN_PASSWORD):
    try:
        email_input = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//input[@id="email"]'))
        )
        password_input = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//input[@id="password"]'))
        )
        login_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//button[@type="submit"]'))
        )
        email_input.send_keys(LOGIN_EMAIL)
        password_input.send_keys(LOGIN_PASSWORD)
        login_button.click()
        print("Logged in successfully.")
    except Exception as e:
        print(f"Error logging in: {e}")
        driver.quit()

# Finding and clicking the add button
def get_to_item_add_page(driver):
    try:
        add_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//a[@class="btn-success btn no-decoration"]'))
        )
        add_button.click()
        driver.refresh() # Refreshing the page
    except Exception as e:
        print(f'Error finding add button: {e}')

# Finding and clicking the add new button
def add_new_item(driver):
    try:
        add_new_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//a[@href="#addnewinventory/"]'))
        )
        add_new_button.click()
        driver.refresh()
    except Exception as e:
        print(f'Error finding add new button: {e}')

# Get products data from Excel
def get_products_data(rows):
    wb = load_workbook(os.getenv('FILE_PATH'), data_only=True)
    ws = wb.active

    data = []
    columns = ws.iter_cols(min_col=1, max_col=16, max_row=rows, values_only=True)

    for i, column in enumerate(columns):
        if i in [1, 3, 4, 5, 8, 13]:
            data.append([cell for cell in column])

    # Splitting the data into individual lists
    name = data[0]
    sku = data[1]
    label = data[2]
    manufacturer = data[3]
    quantity = data[4]
    cost = [round(c, 2) for c in data[5]]
    # print(name, sku, label, manufacturer, quantity, cost)
    return name, sku, label, manufacturer, quantity, cost

# Adding new item
def adding_item(driver, rows):
    name, sku, label, manufacturer, quantity, cost = get_products_data(rows)

    for name, sku, label, manufacturer, quantity, cost in zip(name, sku, label, manufacturer, quantity, cost):
        def new_item_name(driver):
            try:
                item_name = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, '//input[@id="item_name"]'))
                )
                item_name.clear()
                item_name.send_keys(name)
                time.sleep(0.5) # Waiting for 0.5 seconds
            except Exception as e:
                print("Error adding item name:", e)
        
        def new_item_sku(driver):
            try:
                item_sku = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, '//input[@id="item_sku"]'))
                )
                item_sku.clear()
                item_sku.send_keys(f'{sku}-3C')
                time.sleep(0.5)
            except Exception as e:
                print("Error adding item sku:", e)
        
        def new_item_label(driver):
            try:
                item_label = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, '//input[@value="select existing label(s)"]'))
                )
                item_label.click()
                time.sleep(1)

                select_label = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, f'//li[contains(text(), "{label}")]'))
                )
                select_label.click()
                time.sleep(0.5) 
            except Exception as e:
                print("Error adding item label:", e)
        
        def new_item_manufacturer(driver):
            try:
                item_manufacturer = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, '//select[@id="item_manufacturer_id"]'))
                )
                item_manufacturer.click()
                time.sleep(1)

                select_manufacturer = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, f'//option[contains(text(), "{manufacturer}")]'))
                )
                select_manufacturer.click()
                time.sleep(0.5)
            except Exception as e:
                print("Error adding item manufacturer:", e)

        def new_item_date(driver):
            try:
                # Opening the date picker
                item_date_button = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, '//span[@class="add-on"]'))
                )
                item_date_button.click()
                time.sleep(1) 

                # Navigating to the previous month twice
                for _ in range(2):
                    item_date_prev = WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.XPATH, '//th[@class="prev"]'))
                    )
                    item_date_prev.click()
                    time.sleep(0.5)

                # Selecting the exact day (28) from the currently visible month
                item_date_select = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, '//td[@class="day " and text()="28"]'))
                )
                item_date_select.click()
                time.sleep(0.5)  # Waiting for 0.5 seconds 
            except Exception as e:
                print("Error adding item date:", e)
        
        def new_item_quantity(driver):
            try:
                item_quantity = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, '//input[@id="inventory_log_quantity"]'))
                )
                item_quantity.clear()
                item_quantity.send_keys(quantity)
                time.sleep(0.5)
            except Exception as e:
                print("Error adding item quantity:", e)

        def new_item_cost(driver):
            try:
                item_cost = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, '//input[@id="inventory_log_unit_cost"]'))
                )
                item_cost.clear()
                item_cost.send_keys(cost)
                time.sleep(0.5)
            except Exception as e:
                print("Error adding item cost:", e)
        
        # CLicking the save and add another button
        def save_and_next(driver):
            try:
                save_and_next_button = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, '//button[@id="updateSubmitAndAdd"]'))
                )
                save_and_next_button.click()
                time.sleep(0.5)
            except Exception as e:
                print("Error saving and going to next item:", e)

        new_item_name(driver)
        new_item_sku(driver)
        new_item_label(driver)
        new_item_manufacturer(driver)
        new_item_date(driver)
        new_item_quantity(driver)
        new_item_cost(driver)
        save_and_next(driver)
        time.sleep(1)  # Waiting for 1 second

# Holding the page for 20 seconds before quitting      
def hold_page(driver):
    time.sleep(20)
    driver.quit()

def main():
    driver, LOGIN_EMAIL, LOGIN_PASSWORD = initializing()
    login(driver, LOGIN_EMAIL, LOGIN_PASSWORD)
    get_to_item_add_page(driver)
    add_new_item(driver)
    adding_item(driver, rows=37)
    hold_page(driver)

if __name__ == '__main__':
    main()
